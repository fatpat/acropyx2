import logging
from tempfile import NamedTemporaryFile
from openpyxl import Workbook

from models.competitions import Competition, CompetitionResultsExport, CompetitionType
from models.results import RunResultsExport

log = logging.getLogger(__name__)

class CompCtrl:
    @staticmethod
    def start():
        Competition.createIndexes()

    @staticmethod
    def comp_to_xlsx(comp:CompetitionResultsExport, type:CompetitionType):
        ret = None
        wb = Workbook()
        ws = wb.active
        ws.title = f"overall {type}"

        if type == CompetitionType.solo:
            ws.cell(column=1, row=1, value="Rank")
            ws.cell(column=2, row=1, value="Number")
            ws.cell(column=3, row=1, value="First Name")
            ws.cell(column=4, row=1, value="Last Name")
            ws.cell(column=5, row=1, value="Nat")
            ws.cell(column=6, row=1, value="Gender")
            ws.cell(column=7, row=1, value="Glider")
            ws.cell(column=8, row=1, value="Sponsor")
            ws.cell(column=9, row=1, value="CIVL ID")
            ws.cell(column=10, row=1, value="Score")

        if type == CompetitionType.synchro:
            ws.cell(column=1, row=1, value="Rank")
            ws.cell(column=2, row=1, value="Team")
            ws.cell(column=3, row=1, value="Number")
            ws.cell(column=4, row=1, value="First Name")
            ws.cell(column=5, row=1, value="Last Name")
            ws.cell(column=6, row=1, value="Nat")
            ws.cell(column=7, row=1, value="Gender")
            ws.cell(column=8, row=1, value="Glider")
            ws.cell(column=9, row=1, value="Sponsor")
            ws.cell(column=10, row=1, value="CIVL ID")
            ws.cell(column=11, row=1, value="Score")

        rank = 0
        for res in comp.overall_results:
            rank += 1
            score = round(res.score, 3)
            if type == CompetitionType.solo:
                ws.cell(column=1, row=rank+1, value=f"{rank}")
                ws.cell(column=2, row=rank+1, value=f"{res.pilot.civlid}")
                ws.cell(column=3, row=rank+1, value=res.pilot.name.split(" ", 1)[0])
                ws.cell(column=4, row=rank+1, value=res.pilot.name.split(" ", 1)[1])
                ws.cell(column=5, row=rank+1, value=res.pilot.country.upper())
                ws.cell(column=6, row=rank+1, value=f"M")
                ws.cell(column=7, row=rank+1, value=f"")
                ws.cell(column=8, row=rank+1, value=f"")
                ws.cell(column=9, row=rank+1, value=f"{res.pilot.civlid}")
                ws.cell(column=10, row=rank+1, value=f"{score}")

            if type == CompetitionType.synchro:
                for i, pilot in enumerate(res.team.pilots):
                    ws.cell(column=1, row=rank+1+i, value=f"{rank}")
                    ws.cell(column=2, row=rank+1+i, value=f"{res.team.name}")
                    ws.cell(column=3, row=rank+1+i, value=f"{pilot.civlid}")
                    ws.cell(column=4, row=rank+1+i, value=pilot.name.split(" ", 1)[0])
                    ws.cell(column=5, row=rank+1+i, value=pilot.name.split(" ", 1)[1])
                    ws.cell(column=6, row=rank+1+i, value=pilot.country.upper())
                    ws.cell(column=7, row=rank+1+i, value=f"M")
                    ws.cell(column=8, row=rank+1+i, value=f"")
                    ws.cell(column=9, row=rank+1+i, value=f"")
                    ws.cell(column=10, row=rank+1+i, value=f"{pilot.civlid}")
                    ws.cell(column=11, row=rank+1+i, value=f"{score}")
                ws.merge_cells(start_row=rank+1, start_column=1, end_row=rank+2, end_column=1)
                ws.merge_cells(start_row=rank+1, start_column=2, end_row=rank+2, end_column=2)

                rank += 1

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as xlsx:
            ret = xlsx.name
            wb.save(filename=ret)

        return ret

    @staticmethod
    def run_to_xlsx(run:RunResultsExport, type:CompetitionType):
        if not run.final:
            raise HTTPException(400, f"Can't export run because it's not marked as final")
        ret = None
        wb = Workbook()
        ws = wb.active
        ws.title = f"run {type}"

        if type == CompetitionType.solo:
            ws.cell(column=1, row=1, value="Rank")
            ws.cell(column=2, row=1, value="Number")
            ws.cell(column=3, row=1, value="First Name")
            ws.cell(column=4, row=1, value="Last Name")
            ws.cell(column=5, row=1, value="Nat")
            ws.cell(column=6, row=1, value="Gender")
            ws.cell(column=7, row=1, value="Glider")
            ws.cell(column=8, row=1, value="Sponsor")
            ws.cell(column=9, row=1, value="CIVL ID")
            ws.cell(column=10, row=1, value="Score")

        if type == CompetitionType.synchro:
            ws.cell(column=1, row=1, value="Rank")
            ws.cell(column=2, row=1, value="Team")
            ws.cell(column=3, row=1, value="Number")
            ws.cell(column=4, row=1, value="First Name")
            ws.cell(column=5, row=1, value="Last Name")
            ws.cell(column=6, row=1, value="Nat")
            ws.cell(column=7, row=1, value="Gender")
            ws.cell(column=8, row=1, value="Glider")
            ws.cell(column=9, row=1, value="Sponsor")
            ws.cell(column=10, row=1, value="CIVL ID")
            ws.cell(column=11, row=1, value="Score")

        rank = 0
        run.results.sort(key=lambda e: -e.final_marks.score)
        for res in run.results:
            rank += 1
            score = round(res.final_marks.score, 3)
            if type == CompetitionType.solo:
                ws.cell(column=1, row=rank+1, value=f"{rank}")
                ws.cell(column=2, row=rank+1, value=f"{res.pilot.civlid}")
                ws.cell(column=3, row=rank+1, value=res.pilot.name.split(" ", 1)[0])
                ws.cell(column=4, row=rank+1, value=res.pilot.name.split(" ", 1)[1])
                ws.cell(column=5, row=rank+1, value=res.pilot.country.upper())
                ws.cell(column=6, row=rank+1, value=f"M")
                ws.cell(column=7, row=rank+1, value=f"")
                ws.cell(column=8, row=rank+1, value=f"")
                ws.cell(column=9, row=rank+1, value=f"{res.pilot.civlid}")
                ws.cell(column=10, row=rank+1, value=f"{score}")

            if type == CompetitionType.synchro:
                for i, pilot in enumerate(res.team.pilots):
                    ws.cell(column=1, row=rank+1+i, value=f"{rank}")
                    ws.cell(column=2, row=rank+1+i, value=f"{res.team.name}")
                    ws.cell(column=3, row=rank+1+i, value=f"{pilot.civlid}")
                    ws.cell(column=4, row=rank+1+i, value=pilot.name.split(" ", 1)[0])
                    ws.cell(column=5, row=rank+1+i, value=pilot.name.split(" ", 1)[1])
                    ws.cell(column=6, row=rank+1+i, value=pilot.country.upper())
                    ws.cell(column=7, row=rank+1+i, value=f"M")
                    ws.cell(column=8, row=rank+1+i, value=f"")
                    ws.cell(column=9, row=rank+1+i, value=f"")
                    ws.cell(column=10, row=rank+1+i, value=f"{pilot.civlid}")
                    ws.cell(column=11, row=rank+1+i, value=f"{score}")
                ws.merge_cells(start_row=rank+1, start_column=1, end_row=rank+2, end_column=1)
                ws.merge_cells(start_row=rank+1, start_column=2, end_row=rank+2, end_column=2)

                rank += 1

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as xlsx:
            ret = xlsx.name
            wb.save(filename=ret)

        return ret
